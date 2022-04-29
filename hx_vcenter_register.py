#!/usr/bin/env python3

import requests
import sys
import re
import argparse
from openpyxl import load_workbook
from colorama import Fore, Back, Style
import getpass
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

####################################################################################################
#
# COMMAND HELP
#
####################################################################################################

if len(sys.argv) != 3:
    print ('''
    USAGE: hx_vcenter_register.py -f <excel_data_file.xlsx>
    ''')
    sys.exit(0)

####################################################################################################
#
# GET INPUT EXCEL AS ARGUMENT
#
####################################################################################################

parser = argparse.ArgumentParser()
help_str = 'EXCEL file with HyperFlex cluster data.'
parser.add_argument('-f', '--input_file')
args = parser.parse_args()

####################################################################################################
#
# LOAD INPUT EXCEL FILE
#
####################################################################################################


wb = load_workbook(args.input_file)
ws = wb['hx_cluster_info']


####################################################################################################
#
# GET USERNAME AND PASSWORDS IF NECESSARY
#
####################################################################################################


print("")
hx_user = input("Enter HyperFlex username: ")
hx_pass = getpass.getpass("Enter HyperFlex password: ")
vc_user = input("Enter vCenter SSO username: ")
vc_pass = getpass.getpass("Enter vCenter SSO password: ")


####################################################################################################
#
# ITERATE OVER EACH LINE IN EXCEL FILE
#
####################################################################################################

print ("")
for row in ws.iter_rows(min_row=2, values_only=True):

    hx_ip = row[0]
    vc_ip = row[1]

    print ("--------------------------------------------------------------------------------------------")
    print ("HyperFlex Cluster FQDN/IP: "+Style.BRIGHT+hx_ip+Style.RESET_ALL)

    ####################################################################################################
    #
    # GET HYPERFLEX API TOKEN
    #
    ####################################################################################################

    token_url = "https://"+hx_ip+"/aaa/v1/auth?grant_type=password"
    token_data = "{\"username\":\""+hx_user+"\",\"password\":\""+hx_pass+"\",\"client_id\":\"HxGuiClient\",\"client_secret\":\"Sunnyvale\",\"redirect_uri\":\"http://localhost:8080/aaa/redirect\"}"
    token_headers = {"Content-Type": "application/json"}

    #print (token_url)
    #print (token_data)

    try:
        token_response = requests.post(token_url, data = token_data, headers = token_headers, verify = False)
        token_succeed = True
    except:
        token_succeed = False

    if not token_response.ok or token_succeed == False:
        print ("HyperFlex Cluster UUID: unknown")
        print ("HyperFlex Cluster Version: unknown")
        print ("vCenter Register API Supported: unknown")
        print ("vCenter Registration Status: "+Fore.RED+"Not Registered"+Style.RESET_ALL)
        print ("Note: Failed to retrieve HyperFlex API token, check provided cluster information, credentials and/or connectivity")
        print ("")
        continue

    token_response_json = token_response.json()
    token = token_response_json['access_token']

    ####################################################################################################
    #
    # GET HYPERFLEX CLUSTER UUID
    #
    ####################################################################################################

    cuuid_url = "https://"+hx_ip+"/coreapi/v1/clusters"
    cuuid_headers = {"Content-Type": "application/json","Authorization": "Bearer "+token+""}

    try:
        cuuid_response = requests.get(cuuid_url, headers = cuuid_headers, verify = False)
        ccuid_response_json = cuuid_response.json()
        ccuid = ccuid_response_json[0]['uuid']
        print ("HyperFlex Cluster UUID: "+ccuid)
        cuuid_succeed = True
    except:
        cuuid_succeed = False

    if not cuuid_response.ok or cuuid_succeed == False:
        print ("HyperFlex Cluster UUID: unknown")
        print ("HyperFlex Cluster Version: unknown")
        print ("vCenter Register API Supported: unknown")
        print ("vCenter Registration Status: "+Fore.RED+"Not Registered"+Style.RESET_ALL)
        print ("Note: Failed to retrieve HyperFlex UUID, check provided cluster information, credentials and/or connectivity")
        print ("")
        continue

    ####################################################################################################
    #
    # CHECK HYPERFLEX VERSION, MAKE SURE REGISTER API IS SUPPORTED (support starts in HXDP 4.5)
    #
    ####################################################################################################

    version_url = "https://"+hx_ip+"/coreapi/v1/clusters/"+ccuid+"/about"
    version_headers = {"Content-Type": "application/json","Authorization": "Bearer "+token+""}

    try:
        version_response = requests.get(version_url, headers = version_headers, verify = False)
        version_response_json = version_response.json()
        version_full = version_response_json['productVersion']
        version = (version_full.split("-"))[0]
        print ("HyperFlex Cluster Version: "+version)
        version_succeed = True
    except:
        version_succeed = False

    if not cuuid_response.ok or version_succeed == False:
        print ("HyperFlex Cluster Version: unknown")
        print ("vCenter Register API Supported: unknown")
        print ("vCenter Registration Status: "+Fore.RED+"Not Registered"+Style.RESET_ALL)
        print ("Note: Failed to retrieve HyperFlex UUID, check provided cluster information, credentials and/or connectivity")
        print ("")
        continue

    short_version = re.sub(r'.\d[a-z]', '', version)
    if float(short_version) >= float("4.5"):

        print ("vCenter Register API Supported: Yes")

        ####################################################################################################
        #
        # GET VMWARE DATACENTER AND ESXI CLUSTER DETAILS
        #
        ####################################################################################################

        details_url = "https://"+hx_ip+"/coreapi/v1/hypervisor/vcenter"
        details_headers = {"Content-Type":"application/json","Authorization": "Bearer "+token+""}

        try:
            details_response = requests.get(details_url, headers = version_headers, verify = False)
            details_response_json = details_response.json()
            vc_dc = details_response_json['vCenterDatacenter']
            vc_clust = details_response_json['vCenterClusterName']
            print ("VMware Datacenter: "+vc_dc)
            print ("VMware ESXi Cluster: "+vc_clust)
            details_succeed = True
        except:
            details_succeed = False

        if not details_response.ok or details_succeed == False:
            print ("vCenter Registration Status: "+Fore.RED+"Not Registered"+Style.RESET_ALL)
            print ("Note: Failed to retrieve associated VMware Datacenter and ESXi Cluster")
            print ("")
            continue

        ####################################################################################################
        #
        # REGISTER HYPERFLEX CLUSTER WITH VCENTER SERVER
        #
        ####################################################################################################

        register_url = "https://"+hx_ip+"/securityservice/v1/registerNewVC"
        register_data = "{\"newVCenterClustername\":\""+vc_clust+"\",\"newVCenterDatacenter\":\""+vc_dc+"\",\"newVCenterUrl\":\""+vc_ip+"\",\"newVCenterUsername\":\""+vc_user+"\",\"newVCenterPassword\":\""+vc_pass+"\"}"
        register_headers = {"Content-Type":"application/json","Authorization": "Bearer "+token+""}

        try:
            register_response = requests.post(register_url, data = register_data, headers = register_headers, verify = False)
            print ("vCenter Registration Status: "+Fore.GREEN+"Registered"+Style.RESET_ALL)
            print ("Note: None")
            print ("")
            register_succeed = True
        except:
            register_succeed = False

        if not register_response.ok or register_succeed == False:
            print ("vCenter Registration Status: "+Fore.RED+"Not Registered"+Style.RESET_ALL)
            print ("Note: Failed to retrieve HyperFlex UUID, check provided cluster information, credentials and/or connectivity")
            print ("")
            continue

    else:
        print ("vCenter Register API Supported: No")
        print ("vCenter Registration Status: "+Fore.RED+"Not Registered"+Style.RESET_ALL)
        print ("Note: HyperFlex software version does not support vCenter registration API, must use \"stcli cluster reregister\" command")
        print ("")